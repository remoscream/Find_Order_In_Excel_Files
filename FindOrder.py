# coding: utf-8

import os
import sys
import glob
import time
from openpyxl import load_workbook


# Get current folder path
folderpath = os.path.split(os.path.realpath(__file__))[0]

# Initialize file counter
file_counter = 0

# Get filelist
filelist = glob.glob(folderpath + "/*.xlsx")

# Searching
if len(filelist) == 0:
    print('Error ! No order files exist in the folder !')

else:
    # Get order name from a prompt input
    # string_need = input("Enter the order you want to find: \n")

    # Get order name from argumentq
    string_need = sys.argv[1]

    # Load xlsx file
    for f in glob.glob(folderpath + "/*.xlsx"):
        data = load_workbook(f, read_only=True, keep_links=False)

        # Get sheet names in xlsx file
        sheetnames = data.sheetnames

        for m in range(0, len(sheetnames)):
            sheet_now = data[sheetnames[m]]

            for n in range(1, sheet_now.max_row+1):
                celldata = sheet_now.cell(n, 1).value   # Read data of each cell in first column

                if celldata is not None and string_need in celldata:

                    time_file_created = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getmtime(f)))

                    print('row', n, 'in file :', os.path.split(f)[1], '(sheet:', sheetnames[m],
                          '), Order Name:', celldata, ', Created at', time_file_created)

                    file_counter = file_counter + 1

    if file_counter == 0:
        print('Order name:', string_need, 'not found in the files !')



